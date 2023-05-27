import openpyxl
from openpyxl.styles import PatternFill


def adicionar_dados_excel(arquivo_excel, volume_de_dados, tempo, url):
    # Carregar o arquivo Excel existente ou criar um novo se não existir
    try:
        workbook = openpyxl.load_workbook(arquivo_excel)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    # Obter a última linha existente no arquivo Excel
    last_row = sheet.max_row

    # Estilizar a coluna "TEMPO"
    nome_coluna = "TEMPO"
    coluna_nome = sheet.cell(row=1, column=1)
    coluna_nome.value = nome_coluna
    coluna_nome.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")

    # Estilizar a coluna "URL"
    nome_coluna_url = "URL"
    coluna_url = sheet.cell(row=1, column=2)
    coluna_url.value = nome_coluna_url
    coluna_url.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")

        # Fazer a requisição e obter os dados de tempo e URL
        # Inserir os dados nas células
    sheet[f"A{last_row+1}"] = tempo
    sheet[f"B{last_row+1}"] = url

    # Salvar as alterações no arquivo Excel
    workbook.save(arquivo_excel)

